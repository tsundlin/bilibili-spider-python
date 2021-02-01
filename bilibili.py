import time
import openpyxl

from selenium import webdriver
from selenium.webdriver.chrome.webdriver import WebDriver

# 先安装chromedriver 添加入PATH
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait

login_url = "https://passport.bilibili.com/login"
web_url = "https://www.bilibili.com/"
moments_url = "https://t.bilibili.com/?tab=8"
max_video = 20


def create_driver():
    driver = webdriver.Chrome()
    return driver


def open_url(driver: WebDriver):
    driver.get(login_url)

def get_cookie(driver: WebDriver):


def wait_for_login(driver: WebDriver):

    while True:
        if driver.current_url == web_url:
            break
        time.sleep(1)


def collect_moments(driver: WebDriver):
    driver.get(moments_url)
    while True:
        js = 'window.scrollBy(0,10000)'
        driver.execute_script(js)
        videos = driver.find_elements_by_class_name("card")
        if len(videos) >= max_video:
            break
        time.sleep(1)

    return videos


def create_excel(videos):
    """
    """

    # 新建表格
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.cell(1, 1, "up主")
    worksheet.cell(1, 2, "视频")
    worksheet.cell(1, 3, "几分钟前")
    worksheet.cell(1, 4), "简介"

    # excel 起始行数
    row_num = 2
    for video in videos:
        text = video.text.split('\n')

        owner = text[0]
        index = text.index("投稿视频")
        title = text[index + 1]
        time_ago = text[1]
        introduce = text[index + 2]

        worksheet.cell(row_num, 1, owner)
        worksheet.cell(row_num, 2, title)
        worksheet.cell(row_num, 3, time_ago)
        worksheet.cell(row_num, 3, introduce)
        workbook.save('bilibili动态' + '.xlsx')
        row_num = worksheet.max_row + 1


if __name__ == '__main__':
    driver1 = create_driver()
    open_url(driver1)
    wait_for_login(driver1)
    data = collect_moments(driver1)
    create_excel(data)
    driver1.close()
